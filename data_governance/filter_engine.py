"""后置筛选模块 — 读取 raw_manifest，年份/引文/速度漏斗，输出 final_output.xlsx"""

import logging
from pathlib import Path
from dataclasses import dataclass
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rich.console import Console
from rich.panel import Panel
from rich import box

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("filter_engine")

BASE = Path(__file__).resolve().parent.parent
INPUT_PATH = BASE / "xlsx" / "raw_manifest.xlsx"
OUTPUT_PATH = BASE / "xlsx" / "final_output.xlsx"

CURRENT_YEAR = 2026
VELOCITY_LINE = 8.0


@dataclass
class FilterConfig:
    year_min: Optional[int] = None
    year_max: Optional[int] = None
    citation_min: Optional[int] = None
    velocity_min: float = VELOCITY_LINE


def calc_velocity(citations, year) -> float:
    try:
        c = float(citations) if pd.notna(citations) else 0.0
        y = int(year) if pd.notna(year) else CURRENT_YEAR
        if y < 1900 or y > CURRENT_YEAR:
            y = CURRENT_YEAR
        return round(c / max(CURRENT_YEAR - y + 1, 1), 2)
    except (ValueError, TypeError):
        return 0.0


def run(config: Optional[FilterConfig] = None) -> pd.DataFrame:
    if config is None:
        config = FilterConfig()

    console = Console()
    console.print("\n[bold]🔍 后置筛选引擎启动[/bold]\n")

    if not INPUT_PATH.exists():
        logger.error("找不到 %s，请先运行 merge_engine.py", INPUT_PATH)
        return pd.DataFrame()

    df = pd.read_excel(INPUT_PATH, engine="openpyxl")
    before = len(df)
    logger.info("读取 %s → %d 行", INPUT_PATH, before)

    year_col = next((c for c in ["年份", "Year", "year"] if c in df.columns), None)
    cite_col = "引用次数" if "引用次数" in df.columns else None

    df = df.copy()
    df["_velocity"] = 0.0
    df["_filter_result"] = ""

    for idx, row in df.iterrows():
        df.at[idx, "_velocity"] = calc_velocity(
            row.get(cite_col) if cite_col else None,
            row.get(year_col) if year_col else None,
        )

    passed = pd.Series([True] * len(df), index=df.index)

    if year_col and (config.year_min is not None or config.year_max is not None):
        bef = passed.sum()
        if config.year_min is not None:
            passed &= pd.to_numeric(df[year_col], errors="coerce").fillna(0) >= config.year_min
        if config.year_max is not None:
            passed &= pd.to_numeric(df[year_col], errors="coerce").fillna(0) <= config.year_max
        n = bef - passed.sum()
        if n:
            console.print(f"  📅 年份过滤: 剔除 [yellow]{n}[/yellow] 条")

    if config.citation_min is not None and cite_col:
        bef = passed.sum()
        passed &= pd.to_numeric(df[cite_col], errors="coerce").fillna(0) >= config.citation_min
        n = bef - passed.sum()
        if n:
            console.print(f"  📊 引文阈值 ≥ {config.citation_min}: 剔除 [yellow]{n}[/yellow] 条")

    if config.citation_min is not None and config.velocity_min > 0:
        rescued = []
        for idx in df.index:
            if not passed.loc[idx] and df.at[idx, "_velocity"] >= config.velocity_min:
                rescued.append(idx)
                df.at[idx, "_filter_result"] = "🐎 黑马捞回"
        if rescued:
            passed[rescued] = True
            console.print(f"  🐎 黑马捞回: [green]{len(rescued)}[/green] 条 (velocity ≥ {config.velocity_min})")

    result = df[passed].copy()
    result.loc[result["_filter_result"] == "", "_filter_result"] = "✅ 通过"
    after = len(result)
    rescued_count = len(result[result["_filter_result"] == "🐎 黑马捞回"])
    total_cites = int(pd.to_numeric(result[cite_col], errors="coerce").sum() or 0) if cite_col else 0

    yr_range = ""
    if year_col:
        yrs = pd.to_numeric(result[year_col], errors="coerce").dropna()
        if not yrs.empty:
            yr_range = f"{int(yrs.min())}-{int(yrs.max())}"

    lines = [
        "[bold]📋 最终漏斗治理简报[/bold]\n",
        f"总览:     [cyan]{before}[/cyan] → [green]{after}[/green] 条",
        f"年份覆盖: {yr_range}" if yr_range else None,
        f"引文总量: [yellow]{total_cites}[/yellow]",
        f"黑马捞回: [green]{rescued_count}[/green] 条",
        f"速度线:   ≥ {config.velocity_min}/年",
        f"引文阈值: ≥ {config.citation_min}" if config.citation_min is not None else None,
    ]
    console.print(Panel("\n".join(l for l in lines if l), box=box.ROUNDED))

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    out_f = result.copy()
    out_b = df.copy()
    for col in ["_velocity", "_filter_result", "_norm"]:
        for d in (out_f, out_b):
            if col in d.columns:
                d.drop(columns=[col], inplace=True)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        out_f.to_excel(writer, sheet_name="Filtered_High_Value", index=False)
        out_b.to_excel(writer, sheet_name="All_Cleaned_Backup", index=False)

    try:
        wb = load_workbook(OUTPUT_PATH)
        ws = wb["Filtered_High_Value"]
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.fill = green
        wb.save(OUTPUT_PATH)
        wb.close()
    except Exception as e:
        logger.warning("高亮失败: %s", e)

    logger.info("已输出: %s (%d 行)", OUTPUT_PATH, after)
    console.print(f"\n✅ 筛选完成 → [bold]{OUTPUT_PATH}[/bold]\n")
    return result


if __name__ == "__main__":
    run(FilterConfig(year_min=2021, citation_min=50, velocity_min=8))
