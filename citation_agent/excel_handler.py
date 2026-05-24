"""Excel 数据处理器 — 读取/写入 Excel，支持增量保存与断点续传"""
import os
import re
import json
import logging
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from config import (
    EXCEL_PATH,
    PROGRESS_PATH,
    BATCH_SAVE_INTERVAL,
    COLUMN_TITLE,
    COLUMN_CITATIONS,
    COLUMN_CITATION_FMT,
)

logger = logging.getLogger(__name__)


class ExcelHandler:
    """
    Excel 处理器。

    核心设计：
    - 使用 pandas 读取（简洁高效）
    - 使用 openpyxl 写入（原文件基础上按 index 更新，保留格式）
    - 增量保存 + 断点续传
    - 列名通过 config 配置，适应不同格式的 Excel
    """

    def __init__(
        self,
        excel_path: str = None,
        progress_path: str = None,
        col_title: str = None,
        col_citations: str = None,
        col_citation_fmt: str = None,
    ):
        self.excel_path = excel_path or EXCEL_PATH
        self.progress_path = progress_path or PROGRESS_PATH
        self.col_title = col_title or COLUMN_TITLE
        self.col_citations = col_citations or COLUMN_CITATIONS
        self.col_citation_fmt = col_citation_fmt or COLUMN_CITATION_FMT
        self._update_count = 0

    # ─── 读取 ──────────────────────────────────────────

    def load_pending_papers(self) -> pd.DataFrame:
        """
        读取 Excel 并返回待处理的论文行（引用次数为空/NaN）。

        Returns:
            DataFrame 包含所有待处理行，附带原始 DataFrame 索引。
        """
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Excel 文件不存在: {self.excel_path}")

        df = pd.read_excel(self.excel_path, engine="openpyxl")

        if self.col_title not in df.columns:
            raise ValueError(f"Excel 缺少 '{self.col_title}' 列")
        if self.col_citations not in df.columns:
            # 列不存在则创建
            df[self.col_citations] = None

        # 筛选待处理行：引用次数为空或 NaN
        pending = df[df[self.col_citations].isna() | (df[self.col_citations] == "")].copy()

        logger.info(
            "共 %d 条论文，待处理 %d 条",
            len(df),
            len(pending),
        )
        return pending

    def extract_author_from_citation(self, row: pd.Series) -> str:
        """
        从完整学术引用格式中提取第一作者姓名，用于辅助查询。

        格式示例：
        "Ao, S., Hu, Q., Yang, B., ..." → "Ao et al."
        """
        citation_text = row.get(self.col_citation_fmt, "")
        if pd.isna(citation_text) or not str(citation_text).strip():
            return ""

        text = str(citation_text).strip()

        # 匹配引用格式开头的作者名: "LastName, F.," or "LastName, F., & LastName,"
        author_match = re.match(r'^([A-Za-zÀ-ɏ]+)', text)
        if author_match:
            first_author = author_match.group(1)
            # 判断是否有多位作者
            if "&" in text[:100] or ", " in text[:100]:
                return f"{first_author} et al."
            return first_author

        return ""

    def extract_title_from_citation(self, row: pd.Series) -> str:
        """
        从完整学术引用格式中提取论文完整标题。
        """
        citation_text = row.get(self.col_citation_fmt, "")
        short_name = str(row.get(self.col_title, "")).strip()

        if pd.isna(citation_text) or not str(citation_text).strip():
            return short_name

        text = str(citation_text).strip()
        import re

        # 策略 1: ". Title. In:" 模式
        title_match = re.search(
            r'\)[.\s]*([A-Z][A-Za-z0-9\s\-:;,/()\'"]+?)\.\s*(?:In:|In\s|IEEE|Proceedings|Advances|arXiv|bioRxiv|Research|Journal|Conference|Workshop|International|p\.|pp\.)',
            text,
            re.IGNORECASE,
        )
        if title_match:
            extracted = title_match.group(1).strip()
            extracted = re.sub(r'\s+', ' ', extracted).strip().rstrip('.')
            if len(extracted) > 10:
                return extracted

        # 策略 2: 年份后的完整句子
        year_match = re.search(
            r'\b(?:19|20)\d{2}[.\s]+([A-Z][^。]+?)\.\s*(?:In:|In\s|IEEE)', text
        )
        if year_match:
            extracted = year_match.group(1).strip()
            extracted = re.sub(r'\s+', ' ', extracted).strip().rstrip('.')
            if len(extracted) > 10:
                return extracted

        return short_name

    def load_all_papers(self) -> pd.DataFrame:
        """读取全部论文数据"""
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Excel 文件不存在: {self.excel_path}")
        return pd.read_excel(self.excel_path, engine="openpyxl")

    # ─── 增量写入（保留格式） ───────────────────────────

    def update_citation(self, row_index: int, citations: int, status: str = "done") -> None:
        """
        在原 Excel 文件中按行号更新引文数。

        使用 openpyxl 加载工作簿 → 定位单元格 → 写入 → 保存。
        保留原有格式、样式、合并单元格等。

        Args:
            row_index: 原始 DataFrame 中的行号（0-based，不含表头）
            citations: 引文数量
            status: 状态标记（done / error / review / not_found / mismatch）
        """
        wb = load_workbook(self.excel_path)
        ws = wb.active

        # Excel 行号 = row_index + 2（1-based + 表头）
        excel_row = row_index + 2

        # 查找列号
        citations_col = self._find_column(ws, self.col_citations)
        if citations_col is None:
            raise ValueError(f"Excel 中未找到 '{self.col_citations}' 列")

        # 写入引文数
        ws.cell(row=excel_row, column=citations_col, value=citations)

        wb.save(self.excel_path)
        wb.close()

        self._update_count += 1
        logger.debug("已更新 row=%d → citations=%d", row_index, citations)

    def batch_update(
        self,
        updates: list[tuple[int, int, str]],
    ) -> None:
        """
        批量更新多条记录（一次性写入，减少磁盘 I/O）。

        Args:
            updates: [(row_index, citations, status), ...]
        """
        if not updates:
            return

        wb = load_workbook(self.excel_path)
        ws = wb.active

        citations_col = self._find_column(ws, self.col_citations)
        if citations_col is None:
            raise ValueError(f"Excel 中未找到 '{self.col_citations}' 列")

        for row_index, citations, _status in updates:
            excel_row = row_index + 2
            ws.cell(row=excel_row, column=citations_col, value=citations)

        wb.save(self.excel_path)
        wb.close()

        self._update_count += len(updates)
        logger.info("批量更新 %d 条记录", len(updates))

    def _find_column(self, ws, column_name: str) -> Optional[int]:
        """在工作表中查找列名对应的列号（1-based）"""
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and str(header).strip() == column_name:
                return col
        return None

    def should_save(self) -> bool:
        """判断是否达到增量保存阈值"""
        return self._update_count >= BATCH_SAVE_INTERVAL

    def reset_save_counter(self) -> None:
        """重置保存计数器"""
        self._update_count = 0

    def force_save_all(self, results: list[dict]) -> None:
        """
        强制全量写入 Excel（熔断/异常时调用）。

        将当前所有已处理结果批量写回 Excel 文件，
        确保数据不丢失。使用 pandas 全量写回模式。
        """
        try:
            df = pd.read_excel(self.excel_path, engine="openpyxl")
            citations_col = self.col_citations

            if citations_col not in df.columns:
                df[citations_col] = None

            for r in results:
                row_idx = r.get("row_index")
                citations = r.get("citations")
                if row_idx is not None and citations is not None:
                    df.at[row_idx, citations_col] = citations

            df.to_excel(self.excel_path, index=False, engine="openpyxl")
            logger.info("强制全量写入完成: %d 条结果已保存", len(results))
        except Exception as e:
            logger.error("强制全量写入失败: %s", e, exc_info=True)

    # ─── 断点续传 ──────────────────────────────────────

    def save_checkpoint(self, processed: list[dict]) -> None:
        """保存处理进度到 JSON 文件"""
        checkpoint = {
            "processed_count": len(processed),
            "papers": processed,
            "last_updated": str(pd.Timestamp.now()),
        }
        with open(self.progress_path, "w", encoding="utf-8") as f:
            json.dump(checkpoint, f, ensure_ascii=False, indent=2)
        logger.info("检查点已保存: %d 条已处理", len(processed))

    def load_checkpoint(self) -> Optional[list[dict]]:
        """加载断点续传状态"""
        if not os.path.exists(self.progress_path):
            return None
        try:
            with open(self.progress_path, "r", encoding="utf-8") as f:
                checkpoint = json.load(f)
            logger.info(
                "发现检查点: %d 条已处理 (%s)",
                checkpoint.get("processed_count", 0),
                checkpoint.get("last_updated", "unknown"),
            )
            return checkpoint.get("papers", [])
        except (json.JSONDecodeError, KeyError) as e:
            logger.warning("检查点文件损坏: %s", e)
            return None

    def clear_checkpoint(self) -> None:
        """清除检查点"""
        if os.path.exists(self.progress_path):
            os.remove(self.progress_path)
            logger.info("检查点已清除")

    def get_processed_indices(self) -> set[int]:
        """从检查点获取已完成的行号集合"""
        papers = self.load_checkpoint()
        if not papers:
            return set()
        return {p.get("row_index") for p in papers if "row_index" in p}
